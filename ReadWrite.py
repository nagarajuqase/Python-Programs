with open('D:\\RobotFramework_Selenium\\TextRead.txt', 'w+') as f:
    f.write('This is a test\n')
    f.write('5\n')
    value = ('the answer', 42)
    s = str(value)
    f.write(s)
f.close()
f = open('D:\\RobotFramework_Selenium\\TextRead.txt', 'r')
print(f.read())

'''excel = open('D:\\RobotFramework_Selenium\\ReadFile.xlsx', 'r')
print(excel.read())
'''
with open('TextRead2.txt', 'w') as txt:
    txt.write('This is my first text file form directory')
    l = str('\nHHiii')
    txt.write(l)
txt.close()
txt = open('TextRead2.txt', 'r+')
print(txt.read())

import csv
with open('D:\\RobotFramework_Selenium\\CSVFile.csv', 'rb') as csvfile:
    spamreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
    for row in spamreader:
        print ' '.join(row)

with open('D:\\RobotFramework_Selenium\\WriteCSVFile.csv', 'w') as csvfile2:
    fieldnames = ['first_name', 'last_name']
    writer = csv.DictWriter(csvfile2, fieldnames=fieldnames)

    writer.writeheader()
    writer.writerow({'first_name': 'Baked', 'last_name': 'Beans'})
    writer.writerow({'first_name': 'Lovely', 'last_name': 'Spam'})
    writer.writerow({'first_name': 'Wonderful', 'last_name': 'Spam'})

with open('D:\\RobotFramework_Selenium\\WriteCSVFile.csv') as csvfile2:
    reader = csv.DictReader(csvfile2)
    for row in reader:
        print(row['first_name'], row['last_name'])

import openpyxl
'''from openpyxl import load_workbook
wb = load_workbook(filename = 'D:\\RobotFramework_Selenium\\ReadFile.xlsx')
first_sheet = wb.get_sheet_names()[0]
s = wb.get_sheet_by_name(first_sheet)
# print(type(s))
for row in s.iter_rows():
    print row
for cell in row:
    print cell
'''
from openpyxl import load_workbook
import os
wb = load_workbook('D:\\RobotFramework_Selenium\\ReadFile.xlsx')
first_sheet = wb.get_sheet_names()[0]
worksheet = wb.get_sheet_by_name(first_sheet)

for row in worksheet.rows:
    for cell in row:
        print(cell.value),
    print(' ')

ws = wb.create_sheet()
for irow in range(100):
    ws.append(['%d' % i for i in range(200)])
# save the file
wb.save('D:\\RobotFramework_Selenium\\ReadFile.xlsx')
